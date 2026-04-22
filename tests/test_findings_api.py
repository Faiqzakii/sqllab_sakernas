from io import BytesIO
from uuid import uuid4

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlmodel import Session

import app.db as db_module
from app.main import create_app
from app.models import FindingRuleHit


def seed_finding(identity_key: str, nks: str, kode_kab: str) -> None:
    with Session(db_module.engine) as session:
        session.add(
            FindingRuleHit(
                dataset_snapshot_id=1,
                rule_definition_id=1,
                rule_execution_id=1,
                identity_key=identity_key,
                severity="critical",
                message="synthetic issue",
                identity_payload_json={
                    "identity_key": identity_key,
                    "NKS": nks,
                    "KODE_KAB": kode_kab,
                },
            )
        )
        session.commit()


def test_findings_api_filters_by_nks_and_kode_kab() -> None:
    client = TestClient(create_app())
    matching_identity = f"id-{uuid4().hex[:8]}"
    other_identity = f"id-{uuid4().hex[:8]}"
    seed_finding(matching_identity, nks="20250434", kode_kab="71")
    seed_finding(other_identity, nks="20250000", kode_kab="11")

    response = client.get("/api/v1/findings?nks=20250434&kode_kab=71")

    assert response.status_code == 200
    payload = response.json()
    assert [item["identity_key"] for item in payload["data"]] == [matching_identity]
    assert payload["data"][0]["nks"] == "20250434"
    assert payload["data"][0]["kode_kab"] == "71"


def test_findings_excel_export_uses_active_filters() -> None:
    client = TestClient(create_app())
    matching_identity = f"id-{uuid4().hex[:8]}"
    other_identity = f"id-{uuid4().hex[:8]}"
    seed_finding(matching_identity, nks="20250434", kode_kab="71")
    seed_finding(other_identity, nks="20250000", kode_kab="11")

    response = client.get("/api/v1/findings/export.xlsx?kode_kab=71")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(BytesIO(response.content))
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))

    assert rows[0][:4] == ("identity_key", "nks", "kode_kab", "highest_severity")
    assert rows[1][0] == matching_identity
    assert rows[1][2] == "71"
    assert len(rows) == 2
